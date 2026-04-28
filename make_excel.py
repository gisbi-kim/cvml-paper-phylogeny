"""
make_excel.py
=============
papers_classified.json → cvml_taxonomy.xlsx

Sheets:
  1. Papers        — full list with 4-level labels
  2. Taxonomy_Tree — flattened tree (Phylum / Class / Order / Genus / Count)
  3. Stats         — distribution summary
"""
import json
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    raise

IN  = Path("data/intermediate/papers_classified.json")
OUT = Path("cvml_taxonomy.xlsx")

HEADER_FILL  = PatternFill("solid", fgColor="1A73E8")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
PHYLUM_FILL  = PatternFill("solid", fgColor="E8F0FE")
PHYLUM_FONT  = Font(bold=True, color="1A73E8")
CLASS_FILL   = PatternFill("solid", fgColor="F8F9FA")
THIN         = Side(style="thin", color="CCCCCC")
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def auto_width(ws, min_w=8, max_w=60):
    for col in ws.columns:
        length = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

def make_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

def main():
    print(f"Loading {IN} ...")
    with open(IN, encoding="utf-8") as f:
        papers = json.load(f)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Papers ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Papers"
    cols = ["Title", "Venue", "Year", "Authors", "Phylum", "Class", "Order", "Genus", "DOI", "URL"]
    make_header(ws1, cols)

    for i, p in enumerate(papers, 2):
        row = [
            p.get("title",""), p.get("venue",""), p.get("year",""),
            p.get("authors",""),
            p.get("phylum",""), p.get("class",""), p.get("order",""), p.get("genus",""),
            p.get("doi",""), p.get("url",""),
        ]
        for col, val in enumerate(row, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.border = THIN_BORDER
        if i % 10000 == 0:
            print(f"  Sheet 1: {i-1} rows written ...")

    auto_width(ws1)
    print(f"  Sheet 1: {len(papers)} papers done.")

    # ── Sheet 2: Taxonomy_Tree ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Taxonomy_Tree")
    tree_cols = ["Phylum", "Class", "Order", "Genus", "Count"]
    make_header(ws2, tree_cols)

    tree = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
    for p in papers:
        tree[p.get("phylum","Other")][p.get("class","Unclassified")]\
            [p.get("order","Unclassified")][p.get("genus","(general)")] += 1

    row_idx = 2
    for ph, classes in sorted(tree.items()):
        ph_total = sum(sum(sum(g.values()) for g in o.values()) for o in classes.values())
        for cl, orders in sorted(classes.items()):
            cl_total = sum(sum(g.values()) for g in orders.values())
            for od, genera in sorted(orders.items()):
                od_total = sum(genera.values())
                for gn, cnt in sorted(genera.items(), key=lambda x: -x[1]):
                    row_data = [ph, cl, od, gn, cnt]
                    for col, val in enumerate(row_data, 1):
                        cell = ws2.cell(row=row_idx, column=col, value=val)
                        cell.border = THIN_BORDER
                        if col == 1:
                            cell.fill = PHYLUM_FILL
                            cell.font = PHYLUM_FONT
                        elif col == 2:
                            cell.fill = CLASS_FILL
                    row_idx += 1

    auto_width(ws2)
    print(f"  Sheet 2: {row_idx-2} taxonomy rows done.")

    # ── Sheet 3: Stats ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Stats")
    make_header(ws3, ["Phylum", "Papers", "%", "Classes", "Orders"])

    total = len(papers)
    phylum_stats = []
    for ph, classes in tree.items():
        cnt = sum(sum(sum(g.values()) for g in o.values()) for o in classes.values())
        n_classes = len(classes)
        n_orders  = sum(len(orders) for orders in classes.values())
        phylum_stats.append((ph, cnt, n_classes, n_orders))

    for i, (ph, cnt, nc, no) in enumerate(sorted(phylum_stats, key=lambda x: -x[1]), 2):
        pct = cnt / total * 100
        row_data = [ph, cnt, round(pct, 1), nc, no]
        for col, val in enumerate(row_data, 1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.border = THIN_BORDER

    # totals row
    i += 1
    totals = ["TOTAL", total, 100.0,
              sum(x[2] for x in phylum_stats),
              sum(x[3] for x in phylum_stats)]
    for col, val in enumerate(totals, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER

    auto_width(ws3)
    print(f"  Sheet 3: Stats done.")

    wb.save(OUT)
    print(f"\nSaved → {OUT}  ({OUT.stat().st_size/1e6:.1f} MB)")

if __name__ == "__main__":
    main()
